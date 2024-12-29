from openpyxl import Workbook, load_workbook

def create_excel_file(file_name):
    # Create a new workbook
    workbook = Workbook()

    # Select the active worksheet
    sheet1 = workbook.active

    # Set a title for the first worksheet
    sheet1.title = "Sheet1"

    # Add additional worksheets with different names
    # sheet2 = workbook.create_sheet(title="3G Cells_CP")
    # sheet3 = workbook.create_sheet(title="4G Cells_CP")

    # Save the workbook to a file
    workbook.save(file_name)
    print(f"Excel file '{file_name}' has been created successfully!")

file_name = "2G_Daily_Degraded_Cells.xlsx"
create_excel_file(file_name)



def read_and_copy_excel(source_file, destination_file):
    try:
        # Load the source workbook
        source_workbook = load_workbook(source_file)

        # Load or create the destination workbook
        try:
            destination_workbook = load_workbook(destination_file)
        except FileNotFoundError:
            destination_workbook = Workbook()

        # Copy data from each sheet in the source workbook
        for sheet_name in source_workbook.sheetnames:
            source_sheet = source_workbook[sheet_name]
            if sheet_name in destination_workbook.sheetnames:
                destination_sheet = destination_workbook[sheet_name]
            else:
                destination_sheet = destination_workbook.create_sheet(title=sheet_name)

            for row in source_sheet.iter_rows(values_only=True):
                destination_sheet.append(row)

        # Save the destination workbook
        destination_workbook.save(destination_file)
        print(f"Data has been copied from '{source_file}' to '{destination_file}' successfully!")

    except FileNotFoundError:
        print(f"Error: The file '{source_file}' does not exist.")
    except Exception as e:
        print(f"An error occurred: {e}")

# Specify file names
source_file = "2G Cells_Sites_BSC_Region_Cp.xlsx"
destination_file = "2G_Daily_Degraded_Cells.xlsx"

# Call the function to read and copy the Excel file
read_and_copy_excel(source_file, destination_file)


