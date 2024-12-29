import openpyxl

def read_excel_file(file_name):
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(file_name)

        # List all sheet names
        print(f"Sheets in '{file_name}': {workbook.sheetnames}")

        # Read data from each sheet
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"\nData from sheet: {sheet_name}")
            for row in sheet.iter_rows(values_only=True):
                print(row)
    except FileNotFoundError:
        print(f"Error: The file '{file_name}' does not exist.")
    except Exception as e:
        print(f"An error occurred: {e}")

# Specify the file name
file_name = "2G Cells_Sites_BSC_Region_Cp.xlsx"

# Call the function to read the Excel file
read_excel_file(file_name)
